from flask import Flask, request, jsonify
import openpyxl
import os

app = Flask(__name__)

# Nom du fichier Excel
FICHIER_EXCEL = "inscriptions.xlsx"

# Vérifier si le fichier Excel existe, sinon le créer
if not os.path.exists(FICHIER_EXCEL):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Nom", "Prénom", "Classe", "École", "Série", "Adresse", "Téléphone"])  # En-têtes
    wb.save(FICHIER_EXCEL)

@app.route("/ajouter", methods=["POST"])
def ajouter():
    data = request.json

    # Charger le fichier Excel
    wb = openpyxl.load_workbook(FICHIER_EXCEL)
    ws = wb.active

    # Ajouter les données
    ws.append([
        data["nom"], data["prenom"], data["classe"],
        data["ecole"], data["serie"], data["adresse"], data["telephone"]
    ])

    # Sauvegarder le fichier Excel
    wb.save(FICHIER_EXCEL)

    return jsonify({"success": True})

@app.route("/supprimer", methods=["POST"])
def supprimer():
    data = request.json
    nom_a_supprimer = data["nom"]

    # Charger le fichier Excel
    wb = openpyxl.load_workbook(FICHIER_EXCEL)
    ws = wb.active

    # Trouver et supprimer la ligne
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == nom_a_supprimer:  # Comparer avec le nom
            ws.delete_rows(ws.iter_rows().index(row) + 1)
            wb.save(FICHIER_EXCEL)
            return jsonify({"success": True})

    return jsonify({"success": False, "error": "Nom introuvable"})

if __name__ == "__main__":
    app.run(debug=True)
